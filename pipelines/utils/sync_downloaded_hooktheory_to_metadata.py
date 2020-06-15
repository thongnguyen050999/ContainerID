from selenium import webdriver
from bs4 import BeautifulSoup
import pandas as pd
import time
from selenium.webdriver.remote.webelement import WebElement
import time
import json
import os
import shutil
from selenium.common.exceptions import NoSuchElementException
import xlrd

loc = ("hooktheory_metadata.xlsx")
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)
sheet.cell_value(0, 0)



from openpyxl import load_workbook
filepath="hooktheory_metadata.xlsx"
dataset_path = 'C:/Users/Admin/OneDrive - Viralint Pte Ltd/datasets'

w_b=load_workbook(filepath)
sheet_w=w_b.active
arr_json = [x for x in os.listdir(dataset_path) if x.endswith(".json")]
list_check = []
list_genre = []
for ele_arr_json in arr_json:
    _len = len(ele_arr_json)
    _get_name = ele_arr_json[1:_len-6].split('][')
    list_check.append(_get_name[0])
    list_genre.append(_get_name[2])
print (len(list_check))
print (len(list_genre))
col_name_artist = 2
col_section = 3
for row in range(1, sheet.nrows):
    if row == 11053:
        break
    if row > 1 and row <= 11053:
        print ("row: ", row)
        value_url = sheet.cell_value(row, 0)
        value_musicname = sheet.cell_value(row, 1)
        print (value_musicname)
        print ('-------------')
        _string_genre_list = ""
        check_first = 0
        name_author = ""
        check_save = False
        for i in range(len(list_check)):
            if value_musicname == list_check[i]:
                check_save = True
                print (arr_json[i])
                _len = len(arr_json[i])
                _get_name = arr_json[i][1:_len-6].split('][')
                print (_get_name[2])
                if check_first == 0:
                    _string_genre_list += _get_name[2]
                    name_author = _get_name[1]
                else :
                    _string_genre_list += "|" + _get_name[2]
                check_first += 1
        print ("genre: ",_string_genre_list)
        print ("auth: ", name_author)
        if check_save == True:
            sheet_w.cell(row + 1, col_name_artist + 1).value = name_author
            sheet_w.cell(row + 1, col_section + 1).value = _string_genre_list
w_b.save(filepath)