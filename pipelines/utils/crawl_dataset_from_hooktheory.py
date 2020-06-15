from selenium import webdriver
from bs4 import BeautifulSoup
import pandas as pd
import time
from selenium.webdriver.remote.webelement import WebElement
import time
import json
import os
import shutil
from selenium.common.exceptions import NoSuchElementException
import xlrd 
from selenium.webdriver.common.alert import Alert

loc = ("hooktheory_metadata.xlsx")
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)
sheet.cell_value(0, 0)

from openpyxl import load_workbook
filepath="datasmall1.xlsx"
wb=load_workbook(filepath)
sheet_w=wb.active

def check_exists_by_xpath(xpath):
    try:
        webdriver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True

chromeOptions = webdriver.ChromeOptions()
chromeOptions.add_argument("--disable-popup-blocking")
chromedriver = "./chromedriver.exe"
driver = webdriver.Chrome(executable_path=chromedriver, chrome_options=chromeOptions)
driver.get('https://www.hooktheory.com')
# Alert(driver).accept()
# id : vimusicapp
# pass: dustinngo

#id: nhathao95  
#pass: 0913679578Zz
print('Adding cookies successfull')
driver.get('https://www.hooktheory.com')
driver.find_element_by_xpath('//*[@id="bs-navbar-collapse-site"]/ul[2]/li/a').click()
driver.find_element_by_xpath('//*[@id="UserLogin_username"]').send_keys('vimusicapp')
driver.find_element_by_xpath('//*[@id="UserLogin_password"]').send_keys('dustinngo')
driver.find_element_by_xpath('//*[@id="yw0"]/div[3]/button').click()
col_url = 0
col_name_music = 1
col_name_artist = 2
col_section = 3
for row in range(1, sheet.nrows):
    if row >= 8000 and row <= 11053:
        if (sheet.cell_value(row, 2) != ""):
            print ("value:  " + sheet.cell_value(row, 2))
        else:
            print ("empty")
            print ('row: ' + str(row))
            value_url = sheet.cell_value(row, 0)
            print (value_url)
            driver.get(value_url)
            content = driver.page_source
            soup = BeautifulSoup(content)
            name_music = ''
            name_author = ''
            for ele_head in soup.findAll('h1', attrs={'class':'margin-top-0'}):
                name_music = ele_head.text.split('by')[0].strip()
                name_author = ele_head.text.split('by')[1].strip()
            print ("[" +name_music + "][" + name_author + "]")

            print(driver.find_element_by_xpath('/html/body/div[4]/div[1]/div/div/div[1]/div[1]/div[1]/h1'))

            content = driver.page_source
            soup = BeautifulSoup(content)
            last_height = driver.execute_script("return document.body.scrollHeight")
            i = 1
            while i*500 <= last_height:
                print (i)
                driver.execute_script("window.scrollTo(0, "+str(500)+");")
                i += 1
            content = driver.page_source
            soup = BeautifulSoup(content)

            link_list = []
            chor_list = []
            for element in soup.findAll('div', attrs={'class':'btn-group'}):
                chor_list.append(element.find('h2').get('name'))
                link_list.append('https://www.hooktheory.com' + element.find('ul', attrs={'class':'dropdown-menu'}).find('a').get('href'))
            print ('---------')
            n_start = 0
            name_section = ''
            for le_link_list in link_list:
                if n_start == 0:
                    name_section = chor_list[n_start]
                else:
                    name_section += '|' + chor_list[n_start]
                print (le_link_list)
                driver.get(le_link_list)
                time.sleep(5)
                driver.find_element_by_xpath('//*[@id="root"]/div/div[1]/div[1]/div/div[1]/div[3]/div/div/button').click()
                time.sleep(0.5)
                driver.find_element_by_xpath('//*[@id="root"]/div[2]/div/button[5]').click()
                time.sleep(2)
                print( "[" +name_music + "][" + name_author + "][" + chor_list[n_start] + "].json")
                arr_json = [x for x in os.listdir('C:/Users/Admin/Downloads') if x.endswith(".json")]
                for ele_arr_json in arr_json:
                    print (ele_arr_json)
                    oldname = 'C:/Users/Admin/Downloads/' + ele_arr_json
                    newname = 'C:/Users/Admin/Downloads/realData/' + "[" +name_music + "][" + name_author + "][" + chor_list[n_start] + "].json"
                    shutil.move(oldname,newname)
                n_start += 1
            sheet_w.cell(row + 1, col_url + 1).value = value_url
            sheet_w.cell(row + 1, col_name_music + 1).value = name_music
            sheet_w.cell(row + 1, col_name_artist + 1).value = name_author
            sheet_w.cell(row + 1, col_section + 1).value = name_section